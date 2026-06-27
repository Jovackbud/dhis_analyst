"""Generator tests — DOCX, PDF, XLSX, CSV with fixed payloads."""
from backend.app.generators.docx_gen import html_to_docx
from backend.app.generators.pdf_gen import html_to_pdf
from backend.app.generators.xlsx_gen import data_to_xlsx, data_to_csv
from backend.app.generators.pptx_gen import slides_to_pptx


SAMPLE_HTML = """
<h1>Test Report</h1>
<p>This is a test paragraph.</p>
<h2>Data Table</h2>
<table>
<thead><tr><th>Org Unit</th><th>Period</th><th>Value</th></tr></thead>
<tbody>
<tr><td>Kaduna</td><td>2025Q1</td><td>1200</td></tr>
<tr><td>Kaduna</td><td>2025Q2</td><td>1350</td></tr>
</tbody>
</table>
<h2>Recommendations</h2>
<ul>
<li>Validate metadata matches.</li>
<li>Review outliers.</li>
</ul>
"""

SAMPLE_DATA = {
    "rows": [["Kaduna", "2025Q1", 1200], ["Kaduna", "2025Q2", 1350]],
    "headers": ["Org Unit", "Period", "Value"],
    "metadata": {"data_source": "analytics_api"},
    "sources": [{"title": "WHO Malaria Report", "url": "https://who.int/malaria", "confidence": 0.9}],
}

SAMPLE_SLIDES = [
    {"type": "content", "title": "Key Finding", "content": "Malaria cases increased 12%."},
    {"type": "content", "title": "Recommendation", "content": "Investigate outliers."},
]


def test_docx_generates_valid_bytes():
    result = html_to_docx(SAMPLE_HTML, "Test Report")
    assert isinstance(result, bytes)
    assert len(result) > 100
    # DOCX files start with PK (ZIP magic)
    assert result[:2] == b"PK"


def test_pdf_generates_valid_bytes():
    result = html_to_pdf(SAMPLE_HTML)
    assert isinstance(result, bytes)
    assert len(result) > 50
    # PDF files start with %PDF
    assert result[:5] == b"%PDF-"


def test_xlsx_generates_valid_bytes():
    result = data_to_xlsx(SAMPLE_DATA)
    assert isinstance(result, bytes)
    assert len(result) > 100
    # XLSX is also a ZIP
    assert result[:2] == b"PK"


def test_csv_generates_valid_bytes():
    result = data_to_csv(SAMPLE_DATA)
    assert isinstance(result, bytes)
    text = result.decode("utf-8")
    assert "Org Unit" in text
    assert "Kaduna" in text
    assert "1200" in text


def test_pptx_generates_valid_bytes():
    result = slides_to_pptx(SAMPLE_SLIDES, "Test Briefing")
    assert isinstance(result, bytes)
    assert len(result) > 100
    # PPTX is also a ZIP
    assert result[:2] == b"PK"


def test_xlsx_three_sheets():
    from openpyxl import load_workbook
    import io
    result = data_to_xlsx(SAMPLE_DATA)
    wb = load_workbook(io.BytesIO(result))
    assert "Data" in wb.sheetnames
    assert "Metadata" in wb.sheetnames
    assert "Sources" in wb.sheetnames
    # Verify data sheet has rows
    ws = wb["Data"]
    assert ws.max_row >= 3  # header + 2 data rows


def test_docx_empty_html():
    result = html_to_docx("<p></p>", "Empty Report")
    assert isinstance(result, bytes)
    assert result[:2] == b"PK"


def test_csv_empty_data():
    result = data_to_csv({"rows": [], "headers": ["A", "B"]})
    text = result.decode("utf-8")
    assert "A" in text
    assert "B" in text
