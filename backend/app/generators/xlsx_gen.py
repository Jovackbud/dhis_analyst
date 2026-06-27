from __future__ import annotations

import csv
import io
import logging
from typing import Any

from openpyxl import Workbook

logger = logging.getLogger("dhis2_analyst.xlsx_gen")


def data_to_xlsx(payload: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    headers = payload.get("headers") or []
    rows = payload.get("rows") or []
    logger.info("xlsx_generate_start", extra={"row_count": len(rows), "header_count": len(headers)})
    if headers:
        ws.append(headers)
    for row in rows:
        ws.append(row)

    meta = wb.create_sheet("Metadata")
    for key, value in (payload.get("metadata") or {}).items():
        meta.append([key, str(value)])

    sources = wb.create_sheet("Sources")
    sources.append(["title", "url", "confidence"])
    for item in payload.get("sources") or []:
        sources.append([item.get("title", ""), item.get("url", ""), item.get("confidence", "")])

    out = io.BytesIO()
    wb.save(out)
    data = out.getvalue()
    logger.info("xlsx_generate_ok", extra={"size_bytes": len(data), "row_count": len(rows), "sheet_count": len(wb.worksheets)})
    return data


def data_to_csv(payload: dict[str, Any]) -> bytes:
    out = io.StringIO()
    writer = csv.writer(out)
    headers = payload.get("headers") or []
    rows = payload.get("rows") or []
    logger.info("csv_generate_start", extra={"row_count": len(rows), "header_count": len(headers)})
    if headers:
        writer.writerow(headers)
    writer.writerows(rows)
    data = out.getvalue().encode("utf-8")
    logger.info("csv_generate_ok", extra={"size_bytes": len(data), "row_count": len(rows)})
    return data
